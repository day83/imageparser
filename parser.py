from openpyxl import load_workbook
import requests
import lxml
from bs4 import BeautifulSoup
import os
import re
from selenium import webdriver
import time
import sys

'''
Определяет метод работы со скриптами из пользовательского ввода
'''
def get_mode():
    if len(sys.argv) > 1:
        for arg in sys.argv[1:]:
            if arg == '-w' or arg == '--wait-scripts':
                return 'wait-scripts'
    return 'default'

'''
Получает переменную имени сайта из тэга title
'''
def get_site_name(soup):
    site_name_res = soup.select('title')
    if site_name_res:
        site_name = site_name_res[0].text
    else:
        site_name = url
    return site_name

'''
Проверяет существование директории по имени сайта или создаёт таковую.
'''

def create_dir(site_name):
    if os.path.isdir(site_name):
        return True
    try:
        os.mkdir(site_name)
        return True
    except:
        print(f"Error: cannot create folder '{site_name}'")
        return False

'''
Создаёт абсолютную ссылку для скачиваиня изображения в зависимости от формата локальной ссылки.
'''
def generate_image_url(base_url, src):
    # //upload.site.org/...
    if src[:2] == '//':
        url = base_url.split('/')[0] + src
    # /static/...
    elif src[0] == '/':
        url = base_url.split('/')[0] + '//' + base_url.split('/')[2] + '/' + src
    # https://...
    elif src[:4] == 'http':
        url = src
    else:
        url = base_url + '/' + src
    return url

'''
Генерирует имя файла изображения. Добавляет индекс формата [номер] в конец имени в случае,
если такое имя уже существует в директории, но хэши файлов отличаются.
'''
def generate_image_name(site_name, src, compare):
    image_name = src.split('/')[-1]
    while os.path.isfile(site_name + '/' + image_name):
        with open(site_name + '/' + image_name, 'rb') as f:
            data = f.read()
            if hash(data) == compare:
                break
        name, *ext = image_name.rsplit('.', 1)
        if x:=re.search(r" \[(\d+)\]", name):
            index = int(x.group(1)) + 1
            image_name = name.rsplit(' ', 1)[0] + f' [{index}]'
        else:
            image_name = name + f' [1]'
        if ext:
            image_name += '.' + ext[0]
    return image_name

'''
Скачивает изображение по ссылке и сохраняет в файл.
'''
def grab_image(image_url, site_name, src):
    headers = {'User-Agent': 'Mozilla/5.0'}
    print('\t', image_url)
    try:
        res = requests.get(image_url, headers=headers)
        print('\t\t', res.status_code)
        if res.status_code == 200:
            image_name = generate_image_name(site_name, src, hash(res.content))
            try:
                with open(site_name + '/' + image_name, 'wb') as f:
                    f.write(res.content)
            except Exception as e:
                print('\t\t', e)
    except Exception as e:
        print(f"Unable to connect: {image_url}")
        raise e

if __name__ == '__main__':

    mode = get_mode()

    workbook = load_workbook('sites.xlsx')

    sheet = workbook['Book1']

    if mode == 'wait-scripts':
        options = webdriver.FirefoxOptions()
        options.headless = True
        driver = webdriver.Firefox(options=options)

    for i in range(1, sheet.max_row):
        cell = sheet.cell(row=i, column=1).value
        site_url = cell.strip() if cell else None
        if not site_url:
            continue
        if '://' not in site_url:
            site_url = 'https://' + site_url

        try:
            print(f'{site_url}:', end=' ')

            if mode == 'wait-scripts':
                driver.get(site_url)
                time.sleep(5)
                page_data = driver.page_source
            else:
                res = requests.get(site_url)
                page_data = res.text

            print('\tOK')
        except (requests.exceptions.InvalidSchema, requests.exceptions.MissingSchema):
            print('\tNot a valid URL.\n')
            continue
        except Exception:
            print('\tConnection error.\n')
            continue

        soup = BeautifulSoup(page_data, 'lxml')

        site_name = get_site_name(soup)

        if not create_dir(site_name):
            continue

        images = soup.select('img')
        for j, image in enumerate(images):
            src = image.get('src')
            if not src:
                continue
            image_url = generate_image_url(site_url, src)
            res = grab_image(image_url, site_name, src)
        print()
    if mode == 'wait-scripts':
        driver.close()
